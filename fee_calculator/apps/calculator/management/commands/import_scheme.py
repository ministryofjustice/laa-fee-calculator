# -*- coding: utf-8 -*-
from django.core.management.base import BaseCommand

from openpyxl import load_workbook

from calculator.constants import SUTY_BASE_TYPE
from calculator.models import Scheme


def to_list_of_dicts(wb):
    ws = wb.get_active_sheet()
    rows = {}
    header = [c.value.lower() for c in ws[1] if c.value]
    for row in list(ws.rows)[2:]:
        values = []
        for r in row:
            if r.value == 'NULL':
                values.append(None)
            else:
                values.append(r.value)
        if not any(values):
            break
        values_dict = dict(zip(header, values))
        rows[values_dict.pop('fesh_first_id')] = values_dict
    return rows


class Command(BaseCommand):
    help = 'Load Scheme from calculator.xlsx and headers.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '-s', '--calculator', type=str,
            help='xlsx file with calculators in it')
        parser.add_argument(
            '-f', '--header', type=str,
            help='xlsx file with calculator dates in it')

    def handle(self, *args, **options):
        calculator_wb = load_workbook(options['calculator'])
        header_wb = load_workbook(options['header'])

        calculators = to_list_of_dicts(calculator_wb)
        headers = to_list_of_dicts(header_wb)

        for fesh_first_id, calculator in calculators.items():
            header = headers[fesh_first_id]
            s, c = Scheme.objects.get_or_create(
                effective_date=header['effective_date'],
                start_date=header['start_date'],
                end_date=header['end_date'],
                suty_base_type=SUTY_BASE_TYPE.for_constant(
                    calculator['suty_base_type']
                ).value,
                description=calculator['description'],
            )

            if c:
                print('Scheme created %s' % s)
