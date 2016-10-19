# -*- coding: utf-8 -*-
from django.core.management.base import BaseCommand

from openpyxl import load_workbook

from scheme.constants import SUTY_BASE_TYPE
from scheme.models import Scheme


def to_list_of_dicts(wb):
    ws = wb.get_active_sheet()
    rows = {}
    header = [c.value.lower() for c in ws[1] if c.value]
    for row in list(ws.rows)[2:]:
        values = []
        for r in row:
            if r.value == 'NULL':
                values.append(None)
            else:
                values.append(r.value)
        if not any(values):
            break
        values_dict = dict(zip(header, values))
        rows[values_dict.pop('fesh_first_id')] = values_dict
    return rows


class Command(BaseCommand):
    help = 'Load Scheme from scheme.xlsx and headers.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '-s', '--scheme', type=str,
            help='xlsx file with schemes in it')
        parser.add_argument(
            '-f', '--header', type=str,
            help='xlsx file with scheme dates in it')

    def handle(self, *args, **options):
        scheme_wb = load_workbook(options['scheme'])
        header_wb = load_workbook(options['header'])

        schemes = to_list_of_dicts(scheme_wb)
        headers = to_list_of_dicts(header_wb)

        for fesh_first_id, scheme in schemes.items():
            header = headers[fesh_first_id]
            s, c = Scheme.objects.get_or_create(
                effective_date=header['effective_date'],
                start_date=header['start_date'],
                end_date=header['end_date'],
                suty_base_type=SUTY_BASE_TYPE.for_constant(
                    scheme['suty_base_type']
                ).value,
                description=scheme['description'],
            )

            if c:
                print('Scheme created %s' % s)
